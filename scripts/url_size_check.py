#!/usr/bin/env python3
"""
URL 文件大小检测脚本

从 TXT 文件读取 URL 列表，并发获取每个 URL 对应文件的大小（无需下载），
将结果输出为带格式的 Excel 文件。

用法:
    python url_size_check.py <input.txt> [output.xlsx]

参数:
    input.txt    包含 URL 的文本文件，每行一个 URL
    output.xlsx  输出的 Excel 文件路径（可选，默认为 "urlsizecheckresult.xlsx"）

依赖:
    pip install requests openpyxl
"""

import sys
import os
import argparse

try:
    import requests
    from requests.adapters import HTTPAdapter
except ImportError:
    print("错误: 缺少 requests 库，请运行: pip install requests")
    sys.exit(1)

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
except ImportError:
    print("错误: 缺少 openpyxl 库，请运行: pip install openpyxl")
    sys.exit(1)

from concurrent.futures import ThreadPoolExecutor, as_completed


# ── 常量 ──────────────────────────────────────────────

STATUS_MAP = {
    "success": "✓ 成功",
    "no_size": "✗ 无大小信息",
    "timeout": "✗ 超时",
    "conn_error": "✗ 连接失败",
    "req_error": "✗ 请求失败",
    "error": "✗ 错误",
    "failed": "✗ 失败",
}

MAX_WORKERS = 20


# ── 工具函数 ──────────────────────────────────────────

def read_urls_from_txt(file_path):
    """从 TXT 文件读取 URL 列表"""
    urls = []
    with open(file_path, "r", encoding="utf-8") as f:
        for line in f:
            url = line.strip()
            if url and url.startswith(("http://", "https://")):
                urls.append(url)
    return urls


def create_session():
    """创建带连接池的 HTTP Session"""
    s = requests.Session()
    adapter = HTTPAdapter(pool_connections=MAX_WORKERS, pool_maxsize=MAX_WORKERS, max_retries=0)
    s.mount("http://", adapter)
    s.mount("https://", adapter)
    return s


def get_file_size(url, session, retry=1):
    """获取单个 URL 对应文件的大小（字节数）"""
    for attempt in range(retry + 1):
        try:
            # 先尝试 HEAD 请求
            resp = session.head(url, timeout=(3, 5), allow_redirects=True)
            cl = resp.headers.get("Content-Length")
            if cl:
                return int(cl), "success"

            # HEAD 没有 Content-Length，尝试 GET (stream)
            resp = session.get(url, timeout=(3, 5), stream=True, allow_redirects=True)
            cl = resp.headers.get("Content-Length")
            resp.close()
            if cl:
                return int(cl), "success"

            return None, "no_size"
        except requests.exceptions.Timeout:
            if attempt < retry:
                continue
            return None, "timeout"
        except requests.exceptions.ConnectionError:
            if attempt < retry:
                continue
            return None, "conn_error"
        except requests.exceptions.RequestException:
            if attempt < retry:
                continue
            return None, "req_error"
        except Exception:
            if attempt < retry:
                continue
            return None, "error"
    return None, "failed"


def format_size(size):
    """将字节数格式化为可读字符串"""
    if size is None or size == 0:
        return "0 B"
    for unit in ["B", "KB", "MB", "GB", "TB"]:
        if size < 1024:
            if unit == "B":
                return f"{size} B"
            return f"{size:.2f} {unit}"
        size /= 1024
    return f"{size:.2f} PB"


# ── 核心逻辑 ──────────────────────────────────────────

def check_urls(urls):
    """并发检测所有 URL 的文件大小，返回结果列表和汇总"""
    results = []
    session = create_session()
    total = len(urls)

    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        future_to_idx = {
            executor.submit(get_file_size, url, session): (i, url)
            for i, url in enumerate(urls)
        }
        done_count = 0
        for future in as_completed(future_to_idx):
            idx, url = future_to_idx[future]
            size, status = future.result()
            results.append(
                {
                    "index": idx,
                    "url": url,
                    "size": size,
                    "size_formatted": format_size(size) if size else "-",
                    "status": status,
                    "status_text": STATUS_MAP.get(status, "✗ 未知"),
                }
            )
            done_count += 1
            pct = done_count * 100 // total
            print(f"\r  进度: {done_count}/{total} ({pct}%)", end="", flush=True)

    print()  # 换行

    # 按大小降序排列，无大小的排最后
    results.sort(key=lambda r: (r["size"] is None, -(r["size"] or 0)))

    total_size = sum(r["size"] for r in results if r["size"] is not None)
    success_count = sum(1 for r in results if r["size"] is not None)
    fail_count = len(results) - success_count

    summary = {
        "total": len(results),
        "success": success_count,
        "fail": fail_count,
        "total_size": total_size,
        "total_size_formatted": format_size(total_size),
    }

    return results, summary


def write_excel(results, summary, output_path):
    """将检测结果写入带格式的 Excel 文件"""
    wb = Workbook()
    ws = wb.active
    ws.title = "result"

    # 样式
    header_font = Font(bold=True, size=12, color="FFFFFF")
    header_fill = PatternFill(start_color="2563EB", end_color="2563EB", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin", color="E2E8F0"),
        right=Side(style="thin", color="E2E8F0"),
        top=Side(style="thin", color="E2E8F0"),
        bottom=Side(style="thin", color="E2E8F0"),
    )

    # 合计行
    ws.merge_cells("A1:D1")
    total_cell = ws["A1"]
    total_cell.value = (
        f"文件总大小合计: {summary['total_size_formatted']}    "
        f"(共 {summary['total']} 个URL，成功 {summary['success']} 个，"
        f"失败 {summary['fail']} 个)"
    )
    total_cell.font = Font(bold=True, size=14, color="2563EB")
    total_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 36

    # 表头
    headers = ["序号", "URL", "文件大小", "状态"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
    ws.row_dimensions[2].height = 28

    # 数据行
    even_fill = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
    for i, r in enumerate(results):
        row = i + 3
        ws.cell(row=row, column=1, value=i + 1).alignment = Alignment(horizontal="center")
        ws.cell(row=row, column=2, value=r.get("url", ""))
        ws.cell(row=row, column=3, value=r.get("size_formatted", "-")).alignment = Alignment(
            horizontal="right"
        )
        ws.cell(row=row, column=4, value=r.get("status_text", "")).alignment = Alignment(
            horizontal="center"
        )
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.border = thin_border
            if i % 2 == 1:
                cell.fill = even_fill

    # 列宽
    ws.column_dimensions["A"].width = 8
    ws.column_dimensions["B"].width = 70
    ws.column_dimensions["C"].width = 18
    ws.column_dimensions["D"].width = 16

    wb.save(output_path)


# ── 入口 ──────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(
        description="从 TXT 文件读取 URL 列表，检测文件大小并输出 Excel 报告"
    )
    parser.add_argument("input", help="包含 URL 的 TXT 文件路径")
    parser.add_argument(
        "output",
        nargs="?",
        default="urlsizecheckresult.xlsx",
        help="输出 Excel 文件路径（默认: urlsizecheckresult.xlsx）",
    )
    args = parser.parse_args()

    input_path = os.path.abspath(args.input)
    output_path = os.path.abspath(args.output)

    if not os.path.isfile(input_path):
        print(f"错误: 文件不存在 — {input_path}")
        sys.exit(1)

    print(f"📄 读取文件: {input_path}")
    urls = read_urls_from_txt(input_path)

    if not urls:
        print("错误: 文件中未找到有效的 URL（需以 http:// 或 https:// 开头）")
        sys.exit(1)

    print(f"🔗 发现 {len(urls)} 个有效 URL，开始检测...")
    results, summary = check_urls(urls)

    print(f"\n📊 检测完成:")
    print(f"   总数: {summary['total']}")
    print(f"   成功: {summary['success']}")
    print(f"   失败: {summary['fail']}")
    print(f"   总大小: {summary['total_size_formatted']}")

    write_excel(results, summary, output_path)
    print(f"\n✅ 结果已保存: {output_path}")


if __name__ == "__main__":
    main()
